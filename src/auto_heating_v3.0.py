import sys
sys.path.insert(0, "..")
import time
import datetime
from opcua import Client, ua
import openpyxl
import threading
import random
import socket

class style():
    """Style Data"""    
    BLACK = '\033[30m'
    RED = '\033[31m'
    GREEN = '\033[32m'
    YELLOW = '\033[33m'
    BLUE = '\033[34m'
    MAGENTA = '\033[35m'
    CYAN = '\033[36m'
    WHITE = '\033[37m'
    BLACK_BG  = '\33[40m'
    RED_BG    = '\33[41m'
    GREEN_BG  = '\33[42m'
    YELLOW_BG = '\33[43m'
    BLUE_BG   = '\33[44m'
    MAGENTA_BG = '\33[45m'
    CYAN_BG  = '\33[46m'
    WHITE_BG  = '\33[47m'
    UNDERLINE = '\033[4m'
    RESET = '\033[0m'

def print_style(Style_in_ClassStyle, str):
    """Print string with different style"""
    print(Style_in_ClassStyle + str + style.RESET)

def CollectData():
    #idx：Excel欄位指標
    idx_NowTemp = 1 + title.index("NowTemp")
    idx_Time = 1 + title.index("Time")
    idx_Duration = 1 + title.index("Duration")
    idx_isHeat = 1 + title.index("isHeat")
    idx_NextTemp = 1 + title.index("NextTemp")
    max_row = sheet.max_row

    reset_time = datetime.datetime.now()
    now_temp = client.get_node("ns=3;s=\"dbAppCtrl\".\"Hmi\".\"Obj\".\"EB\".\"Proc\".\"rActVal\"").get_value()
    # reset_time = datetime.datetime.now()
    data_count = 0
    while True:
        now_time = datetime.datetime.now()
        if data_count == 100:
            reset_time = now_time
            data_count = 0
            print("完成一循環紀錄(100筆)")
        heating = client.get_node("ns=3;s=\"xEB_KF1\"").get_value()
        time.sleep(0.97)
        next_temp = client.get_node("ns=3;s=\"dbAppCtrl\".\"Hmi\".\"Obj\".\"EB\".\"Proc\".\"rActVal\"").get_value()
        sheet.cell(1 + max_row, idx_NowTemp).value = now_temp
        sheet.cell(1 + max_row, idx_Time).value = now_time
        sheet.cell(1 + max_row, idx_isHeat).value = int(heating)
        sheet.cell(1 + max_row, idx_NextTemp).value = next_temp
        dTime = (now_time - reset_time).total_seconds()
        sheet.cell(1 + max_row, idx_Duration).value = dTime
        data_count += 1
        now_temp = next_temp
        max_row = sheet.max_row
        print("紀錄第 ", data_count, " 筆")

def BeltControl():
    try:
        count = -1
        while True:
            automatic = client.get_node("ns=3;s=\"dbVar\".\"OpMode\".\"Auto\".\"xAct\"")
            bg26 = client.get_node("ns=3;s=\"xG1_BG26\"").get_value()
            if bg26:
                if count == -1:
                    count = 0
                    time.sleep(5)
                else:
                    automatic.set_attribute(ua.AttributeIds.Value, ua.DataValue(False))
                    #等待時間
                    wait_time = int(60 * random.random())
                    print("本次等待時間：", wait_time)
                    time.sleep(wait_time)
                    automatic.set_attribute(ua.AttributeIds.Value, ua.DataValue(True))
                    time.sleep(5)
                    count += 1
            time.sleep(0.1)
    except:
        pass
            
if __name__ == "__main__":
    #讀取Excel檔案
    try:
        date = datetime.datetime.now()
        wb_name = 'HeatingData_' + date.strftime("%m%d") + '.xlsx'
        wb = openpyxl.load_workbook(wb_name)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.worksheets[0]
        sheet.title = "data"
    #讀取工作表
    try:
        sheet = wb["data"]
    except KeyError:
        sheet = wb.create_sheet("data")
    #寫入工作表標題
    title = ["NowTemp","Time","Duration","isHeat","NextTemp"]
    for i in range(len(title)):
        box = sheet.cell(row=1, column=i+1)
        if box.value != title[i]:
            box.value = title[i]
    wb.save(wb_name)
    print("Excel檔案讀取成功")

    # OPCUA連線
    client = Client("opc.tcp://172.21.8.1:4840/")
    try:
        client.connect()
        # 輸送帶節點資訊
        automatic = client.get_node("ns=3;s=\"dbVar\".\"OpMode\".\"Auto\".\"xAct\"")
        # 暫停輸送帶
        # automatic.set_attribute(ua.AttributeIds.Value, ua.DataValue(False))
        
        # 目標加熱溫度
        target_temp = 30
        # while True:
        #     print("請輸入目標加熱溫度(常用30度)")
        #     target_temp = int(input("> "))
        #     if target_temp > 0: break
        #     else: print_style(style.RED,"目標加熱溫度必須大於0")

        # 啟動輸送帶
        automatic.set_attribute(ua.AttributeIds.Value, ua.DataValue(True))
        time.sleep(1)
        collecting = threading.Thread(target = CollectData)
        belt_delay = threading.Thread(target = BeltControl)
        collecting.setDaemon(True)
        belt_delay.setDaemon(True)
        collecting.start()
        belt_delay.start()
        while True:
            cmd = input("")
            if cmd == "exit":
                heating = client.get_node("ns=3;s=\"xEB_KF1\"").get_value()
                if heating:
                    print("尚在加熱中，無法停止程式。請待加熱完成。")
                else:
                    automatic.set_attribute(ua.AttributeIds.Value, ua.DataValue(False))
                    break
            elif cmd == "force exit":
                break
            else:
                print_style(style.RED,"無此指令")
        
        wb.save(wb_name)
        wb.close()
    except socket.error as e:
        print(e)
        print("請確認是否已連上Taiwan-CP-Factory後再啟動程式")
        input("Press 'Enter' to exit...")
    finally:
        client.disconnect()
