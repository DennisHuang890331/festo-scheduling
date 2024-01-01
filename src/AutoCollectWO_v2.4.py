import sys
sys.path.insert(0, "..")
import datetime
from opcua import Client, ua
import socket
import openpyxl
import threading
import collections

class style():
    """Style Data"""    
    BLACK = '\033[30m'
    RED = '\033[31m'
    GREEN = '\033[32m'
    YELLOW = '\033[33m'
    BLUE = '\033[34m'
    MAGENTA = '\033[35m'
    CYAN = '\033[36m'
    WHITE = '\033[37m'
    BLACK_BG  = '\33[40m'
    RED_BG    = '\33[41m'
    GREEN_BG  = '\33[42m'
    YELLOW_BG = '\33[43m'
    BLUE_BG   = '\33[44m'
    MAGENTA_BG = '\33[45m'
    CYAN_BG  = '\33[46m'
    WHITE_BG  = '\33[47m'
    UNDERLINE = '\033[4m'
    RESET = '\033[0m'
def print_style(Style_in_ClassStyle, str):
    """Print string with different style"""
    print(Style_in_ClassStyle + str + style.RESET)

def excel_save(wbname, filename):
    """Excel存檔"""
    try:
        wbname.save(filename)
        print("檔案儲存成功：" + filename)
    except PermissionError:
        print_style(style.RED, "Permission denied: " + filename)
        print_style(style.RED, "請關閉Excel後，輸入'retry'再試一次，或輸入'cancel'結束程式。")
        while True:
            cmd_wbsave = input("> ")
            if cmd_wbsave.lower() == "retry":
                excel_save(wbname, filename)
                break
            elif cmd_wbsave.lower() == "cancel":
                exit()
            else:
                print_style(style.RED, "指令不存在。")

def terminal():
    """多執行緒用函式，用來判斷是否結束程式"""
    while True:
        cmd = input(">> ")
        if cmd.lower() == "exit":
            break
        else:
            print_style(style.RED,"指令不存在")

def main():
    global sh
    try:
        file_name = "tblFinStep.xlsx"
        sheet_name = "tblFinStep"
        wb = openpyxl.load_workbook(file_name)
    except FileNotFoundError: #若找不到檔案則新增檔案
        wb = openpyxl.Workbook() #開新檔案
        sh = wb.worksheets[0] #開新工作表
        sh.title = sheet_name #重新命名工作表
    #FESTO資料庫工作表標題
    db_title = ["WPNo","StepNo","ONo","OPos","Description","OpNo","NextStepNo","FirstStep","ErrorStepNo","NewPNo",\
                "PlanedStart","PlanedEnd","Start","End","OPNoType","ResourceID","TransportTime","ErrorStep",\
                "ElectricEnergyCalc","ElectricEnergyReal","CompressedAirCalc","CompressedAirReal","FreeString",\
                "StaffId","ErrorRetVal","Heat iniTemp.","Heat finTemp."]
    #讀取工作表
    try:
        sh = wb[sheet_name]
        for i in range(len(db_title)):
            box = sh.cell(row=1, column=i+1)
            if box.value == None:
                box.value = db_title[i]
            elif box.value != db_title[i]:
                raise KeyError
    #例外處理
    except KeyError:
        print_style(style.RED, "既有 tblFinStep.xlsx 格式錯誤，請修正後再繼續。")
        exit()
    print("Excel檔案讀取成功")

    #資料庫標題INDEX (給Excel寫入)
    idx_WPNo = db_title.index("WPNo") + 1 #工單類型
    idx_oNo = db_title.index("ONo") + 1 #工單編號
    idx_oPos = db_title.index("OPos") + 1 #該工單中第幾個工件
    idx_Desc = db_title.index("Description") + 1 #加工程序描述
    idx_OpNo = db_title.index("OpNo") + 1 #加工後工件類型
    idx_Start = db_title.index("Start") + 1 #加工開始時間
    idx_end = db_title.index("End") + 1 #加工結束時間
    idx_iniTemp = db_title.index("Heat iniTemp.") + 1 #加熱起始溫度
    idx_finTemp = db_title.index("Heat finTemp.") + 1 #加熱結束溫度
    
    #OPCUA連線IP
    Storage = Client("opc.tcp://172.21.3.1:4840/")
    Measure = Client("opc.tcp://172.21.6.1:4840/")
    Drill = Client("opc.tcp://172.21.7.1:4840/")
    Pcb = Client("opc.tcp://172.21.5.1:4840/")
    Visual = Client("opc.tcp://172.21.4.1:4840/")
    Cover = Client("opc.tcp://172.21.1.1:4840/")
    Press = Client("opc.tcp://172.21.2.1:4840/")
    Heating = Client("opc.tcp://172.21.8.1:4840/")
    try:
        #OPCUA連線
        Storage.connect()
        Measure.connect()
        Drill.connect()
        Pcb.connect()
        Visual.connect()
        Cover.connect()
        Press.connect()
        Heating.connect()
        
        #倉儲節點資訊
        S_mes_busy = Storage.get_node("ns=3;s=\"dbMes\".\"xBusy\"")
        S_bg21_1 = Storage.get_node("ns=3;s=\"xG1_BG21\"")
        S_WPNo_1 = Storage.get_node("ns=3;s=\"dbRfidData\".\"ID1\".\"Mes\".\"diPNo\"")
        S_ONo_1 = Storage.get_node("ns=3;s=\"dbRfidData\".\"ID1\".\"Mes\".\"diONo\"")
        S_oPos_1 = Storage.get_node("ns=3;s=\"dbRfidData\".\"ID1\".\"Mes\".\"iOPos\"")
        S_OpNo_1 = Storage.get_node("ns=3;s=\"dbRfidData\".\"ID1\".\"Mes\".\"iOpNo\"")
        S_Resource_1 = Storage.get_node("ns=3;s=\"dbRfidData\".\"ID1\".\"Mes\".\"iResourceId\"")
        S_CarrierID_1 = Storage.get_node("ns=3;s=\"dbRfidData\".\"ID1\".\"iCarrierID\"")
        #量測節點資訊
        M_mes_busy = Measure.get_node("ns=3;s=\"dbMes\".\"xBusy\"")
        M_bg21 = Measure.get_node("ns=3;s=\"xW1_BG21\"")
        M_bg51_1 = Measure.get_node("ns=3;s=\"xW1_BG51\"")
        M_bg41 = Measure.get_node("ns=3;s=\"xW1_BG41\"")
        M_bg42 = Measure.get_node("ns=3;s=\"xW1_BG42\"")
        M_WPNo_1 = Measure.get_node("ns=3;s=\"dbRfidData\".\"ID2\".\"Mes\".\"diPNo\"")
        M_ONo_1 = Measure.get_node("ns=3;s=\"dbRfidData\".\"ID2\".\"Mes\".\"diONo\"")
        M_oPos_1 = Measure.get_node("ns=3;s=\"dbRfidData\".\"ID2\".\"Mes\".\"iOPos\"")
        M_OpNo_1 = Measure.get_node("ns=3;s=\"dbRfidData\".\"ID2\".\"Mes\".\"iOpNo\"")
        M_Resource_1 = Measure.get_node("ns=3;s=\"dbRfidData\".\"ID2\".\"Mes\".\"iResourceId\"")
        M_CarrierID_1 = Measure.get_node("ns=3;s=\"dbRfidData\".\"ID2\".\"iCarrierID\"")
        #2:準備進AGV
        M_WPNo_2 = Measure.get_node("ns=3;s=\"dbRfidData\".\"ID1\".\"Mes\".\"diPNo\"")
        M_ONo_2 = Measure.get_node("ns=3;s=\"dbRfidData\".\"ID1\".\"Mes\".\"diONo\"")
        M_oPos_2 = Measure.get_node("ns=3;s=\"dbRfidData\".\"ID1\".\"Mes\".\"iOPos\"")
        M_OpNo_2 = Measure.get_node("ns=3;s=\"dbRfidData\".\"ID1\".\"Mes\".\"iOpNo\"")
        M_Resource_2 = Measure.get_node("ns=3;s=\"dbRfidData\".\"ID1\".\"Mes\".\"iResourceId\"")
        M_CarrierID_2 = Measure.get_node("ns=3;s=\"dbRfidData\".\"ID1\".\"iCarrierID\"")
        #鑽孔節點
        D_mes_busy = Drill.get_node("ns=3;s=\"dbMes\".\"xBusy\"")
        D_bg21 = Drill.get_node("ns=3;s=\"xG1_BG21\"")
        D_WPNo = Drill.get_node("ns=3;s=\"dbRfidData\".\"ID1\".\"Mes\".\"diPNo\"")
        D_ONo = Drill.get_node("ns=3;s=\"dbRfidData\".\"ID1\".\"Mes\".\"diONo\"")
        D_oPos = Drill.get_node("ns=3;s=\"dbRfidData\".\"ID1\".\"Mes\".\"iOPos\"")
        D_OpNo = Drill.get_node("ns=3;s=\"dbRfidData\".\"ID1\".\"Mes\".\"iOpNo\"")
        D_Resource = Drill.get_node("ns=3;s=\"dbRfidData\".\"ID1\".\"Mes\".\"iResourceId\"")
        D_CarrierID = Drill.get_node("ns=3;s=\"dbRfidData\".\"ID1\".\"iCarrierID\"")
        #PCB放置節點
        pcb_mes_busy = Pcb.get_node("ns=3;s=\"dbMes\".\"xBusy\"")
        pcb_bg21 = Pcb.get_node("ns=3;s=\"xG1_BG21\"")
        pcb_bg31 = Pcb.get_node("ns=3;s=\"xG1_BG31\"")
        pcb_WPNo = Pcb.get_node("ns=3;s=\"dbRfidData\".\"ID2\".\"Mes\".\"diPNo\"")
        pcb_ONo = Pcb.get_node("ns=3;s=\"dbRfidData\".\"ID2\".\"Mes\".\"diONo\"")
        pcb_oPos = Pcb.get_node("ns=3;s=\"dbRfidData\".\"ID2\".\"Mes\".\"iOPos\"")
        pcb_OpNo = Pcb.get_node("ns=3;s=\"dbRfidData\".\"ID2\".\"Mes\".\"iOpNo\"")
        pcb_Resource = Pcb.get_node("ns=3;s=\"dbRfidData\".\"ID2\".\"Mes\".\"iResourceId\"")
        pcb_CarrierID = Pcb.get_node("ns=3;s=\"dbRfidData\".\"ID2\".\"iCarrierID\"")
        #視覺檢測節點
        V_mes_busy = Visual.get_node("ns=3;s=\"dbMes\".\"xBusy\"")
        V_bg21 = Visual.get_node("ns=3;s=\"xW1_BG21\"")
        V_bg51_1 = Visual.get_node("ns=3;s=\"xW1_BG51\"")
        V_bg42 = Visual.get_node("ns=3;s=\"xW1_BG42\"")
        V_bg41 = Visual.get_node("ns=3;s=\"xW1_BG41\"")
        V_bg31 = Visual.get_node("ns=3;s=\"xW1_BG31\"")
        V_WPNo_1 = Visual.get_node("ns=3;s=\"dbRfidData\".\"ID2\".\"Mes\".\"diPNo\"")
        V_ONo_1 = Visual.get_node("ns=3;s=\"dbRfidData\".\"ID2\".\"Mes\".\"diONo\"")
        V_oPos_1 = Visual.get_node("ns=3;s=\"dbRfidData\".\"ID2\".\"Mes\".\"iOPos\"")
        V_OpNo_1 = Visual.get_node("ns=3;s=\"dbRfidData\".\"ID2\".\"Mes\".\"iOpNo\"")
        V_Resource_1 = Visual.get_node("ns=3;s=\"dbRfidData\".\"ID2\".\"Mes\".\"iResourceId\"")
        V_CarrierID_1 = Visual.get_node("ns=3;s=\"dbRfidData\".\"ID2\".\"iCarrierID\"")
        V_WPNo_2 = Visual.get_node("ns=3;s=\"dbRfidData\".\"ID1\".\"Mes\".\"diPNo\"")
        V_ONo_2 = Visual.get_node("ns=3;s=\"dbRfidData\".\"ID1\".\"Mes\".\"diONo\"")
        V_oPos_2 = Visual.get_node("ns=3;s=\"dbRfidData\".\"ID1\".\"Mes\".\"iOPos\"")
        V_OpNo_2 = Visual.get_node("ns=3;s=\"dbRfidData\".\"ID1\".\"Mes\".\"iOpNo\"")
        V_Resource_2 = Visual.get_node("ns=3;s=\"dbRfidData\".\"ID1\".\"Mes\".\"iResourceId\"")
        V_CarrierID_2 = Visual.get_node("ns=3;s=\"dbRfidData\".\"ID1\".\"iCarrierID\"")
        #加蓋節點
        C_mes_busy = Cover.get_node("ns=3;s=\"dbMes\".\"xBusy\"")
        C_bg21 = Cover.get_node("ns=3;s=\"xG1_BG21\"")
        C_WPNo = Cover.get_node("ns=3;s=\"dbRfidData\".\"ID1\".\"Mes\".\"diPNo\"")
        C_ONo = Cover.get_node("ns=3;s=\"dbRfidData\".\"ID1\".\"Mes\".\"diONo\"")
        C_oPos = Cover.get_node("ns=3;s=\"dbRfidData\".\"ID1\".\"Mes\".\"iOPos\"")
        C_OpNo = Cover.get_node("ns=3;s=\"dbRfidData\".\"ID1\".\"Mes\".\"iOpNo\"")
        C_Resource = Cover.get_node("ns=3;s=\"dbRfidData\".\"ID1\".\"Mes\".\"iResourceId\"")
        C_CarrierID = Cover.get_node("ns=3;s=\"dbRfidData\".\"ID1\".\"iCarrierID\"")
        #加壓節點
        P_mes_busy = Press.get_node("ns=3;s=\"dbMes\".\"xBusy\"")
        P_bg21 = Press.get_node("ns=3;s=\"xG1_BG21\"")
        P_WPNo = Press.get_node("ns=3;s=\"dbRfidData\".\"ID1\".\"Mes\".\"diPNo\"")
        P_ONo = Press.get_node("ns=3;s=\"dbRfidData\".\"ID1\".\"Mes\".\"diONo\"")
        P_oPos = Press.get_node("ns=3;s=\"dbRfidData\".\"ID1\".\"Mes\".\"iOPos\"")
        P_OpNo = Press.get_node("ns=3;s=\"dbRfidData\".\"ID1\".\"Mes\".\"iOpNo\"")
        P_Resource = Press.get_node("ns=3;s=\"dbRfidData\".\"ID1\".\"Mes\".\"iResourceId\"")
        P_CarrierID = Press.get_node("ns=3;s=\"dbRfidData\".\"ID1\".\"iCarrierID\"")
        #加熱節點
        H_manual_heat = Heating.get_node("ns=3;s=\"xEB_KF1\"")
        pwr1000 = Heating.get_node("ns=3;s=\"xEB_KF2\"")
        H_setup_mode = Heating.get_node("ns=3;s=\"dbVar\".\"OpMode\".\"Man\".\"xAct\"")
        H_automatic = Heating.get_node("ns=3;s=\"dbVar\".\"OpMode\".\"Auto\".\"xAct\"")
        H_actual_temp = Heating.get_node("ns=3;s=\"dbAppCtrl\".\"Hmi\".\"Obj\".\"EB\".\"Proc\".\"rActVal\"")
        H_bg26 = Heating.get_node("ns=3;s=\"xG1_BG26\"")
        H_bg21 = Heating.get_node("ns=3;s=\"xG1_BG21\"")
        H_mb20 = Heating.get_node("ns=3;s=\"xG1_MB20\"")
        H_belt_go = Heating.get_node("ns=3;s=\"xQA1_A1\"")
        H_mes_busy = Heating.get_node("ns=3;s=\"dbMes\".\"xBusy\"")
        H_WPNo = Heating.get_node("ns=3;s=\"dbRfidData\".\"ID1\".\"Mes\".\"diPNo\"")
        H_ONo = Heating.get_node("ns=3;s=\"dbRfidData\".\"ID1\".\"Mes\".\"diONo\"")
        H_oPos = Heating.get_node("ns=3;s=\"dbRfidData\".\"ID1\".\"Mes\".\"iOPos\"")
        H_OpNo = Heating.get_node("ns=3;s=\"dbRfidData\".\"ID1\".\"Mes\".\"iOpNo\"")
        H_Resource = Heating.get_node("ns=3;s=\"dbRfidData\".\"ID1\".\"Mes\".\"iResourceId\"")
        H_CarrierID = Heating.get_node("ns=3;s=\"dbRfidData\".\"ID1\".\"iCarrierID\"")
        
        #旗標FOR機台是否正在做事
        isStorage =False
        isMeasure = False
        isDrill = False
        isPcb = False
        isVaual = False
        isCover = False
        isPress = False
        isHeating = False
        isPreheat = False
        isAgvGoWait = False
        isAgvGo = False
        isAgvBackWait = False
        isAgvBack = False
        isRfidReadytoNote_Go = False
        isRfidReadytoNote_Back = False
        
        PreheatStop = False #是否停止預熱
        PreheatStopCheck = False #旗標為了防止

        #RFID資訊儲存格式，用來判斷AGV進出時間與物料資訊
        rfid_data = collections.namedtuple('rfid_data','wpno, ono, opos, opno')
        rfid_agv_go = []
        rfid_agv_back = []
        
        #用來判斷工單是否完成
        max_order_number = 0 #該工單總筆數
        order_number_finished = 0 #已完成工單數

        print_style(style.CYAN,"開始收資料")

        while cmd_terminal.is_alive():
            #加熱站預熱
            if not PreheatStop:
                if H_actual_temp.get_value() < 28 and not isPreheat:
                    isPreheat = True
                    H_automatic.set_attribute(ua.AttributeIds.Value, ua.DataValue(False))
                    H_setup_mode.set_attribute(ua.AttributeIds.Value, ua.DataValue(True))
                    H_mb20.set_attribute(ua.AttributeIds.Value, ua.DataValue(True))
                    H_belt_go.set_attribute(ua.AttributeIds.Value, ua.DataValue(True))
                    H_manual_heat.set_attribute(ua.AttributeIds.Value, ua.DataValue(True))
                if H_actual_temp.get_value() > 28.5:
                    isPreheat = False
                    H_automatic.set_attribute(ua.AttributeIds.Value, ua.DataValue(True))
                if D_bg21.get_value() and not PreheatStopCheck:
                    PreheatStopCheck = True
                elif not D_bg21.get_value() and PreheatStopCheck and D_OpNo.get_value() == 112:
                    PreheatStopCheck = False
                    PreheatStop = True
                    H_automatic.set_attribute(ua.AttributeIds.Value, ua.DataValue(True))
                    print("結束預熱")

            #各加工站加工時間紀錄
            nowtime = datetime.datetime.now()
            #倉儲出料、回料
            if S_mes_busy.get_value() and S_bg21_1.get_value() and not isStorage:
                isStorage = True
                S_row = sh.max_row + 1
                sh.cell(S_row, idx_Start).value = nowtime
                rfid_1 = rfid_data(S_WPNo_1.get_value(),S_ONo_1.get_value(),S_oPos_1.get_value(),S_OpNo_1.get_value())
                print("entered Storage")
                # print("rfid1：",rfid_1)
            elif not S_mes_busy.get_value() and isStorage:
                sh.cell(S_row, idx_end).value = nowtime
            if not S_bg21_1.get_value() and isStorage:
                isStorage = False
                rfid_2 = rfid_data(S_WPNo_1.get_value(),S_ONo_1.get_value(),S_oPos_1.get_value(),S_OpNo_1.get_value())
                # print("rfid2：",rfid_2)
                if rfid_1.ono == 0:
                    sh.cell(S_row, idx_WPNo).value = rfid_2.wpno
                    sh.cell(S_row, idx_oNo).value = rfid_2.ono
                    sh.cell(S_row, idx_oPos).value = rfid_2.opos
                    sh.cell(S_row, idx_OpNo).value = 212
                    sh.cell(S_row, idx_Desc).value = "release a defined part on stopper 1"
                    print("release a defined part on stopper 1")
                else:
                    sh.cell(S_row, idx_WPNo).value = rfid_1.wpno
                    sh.cell(S_row, idx_oNo).value = rfid_1.ono
                    sh.cell(S_row, idx_OpNo).value = rfid_1.opno
                    sh.cell(S_row, idx_oPos).value = rfid_1.opos
                    sh.cell(S_row, idx_Desc).value = "store a part from stopper 1"
                    print("store a part from stopper 1")
                    order_number_finished += 1
                    if order_number_finished == max_order_number:
                        PreheatStop = False
                        max_order_number = 0
                        print_style(style.RED, "本次工單結束")
            # 去程AGV
            if V_bg21.get_value() and not isRfidReadytoNote_Go :
                isRfidReadytoNote_Go = True
                # last_opos_go = rfid_agv_go[-1].opos
            elif not V_bg21.get_value() and isRfidReadytoNote_Go and V_OpNo_2.get_value() in [115, 120, 121, 122, 123, 112]:
                isRfidReadytoNote_Go = False
                rfid_agv_go.append(rfid_data(V_WPNo_2.get_value(),V_ONo_2.get_value(),V_oPos_2.get_value(),V_OpNo_2.get_value()))
                print("AGV Waiting (go)", rfid_agv_go)
            if V_bg42.get_value() and not isAgvGoWait:
                isAgvGoWait = True
            elif not V_bg42.get_value() and isAgvGoWait:
                isAgvGoWait = False
                isAgvGo = True
                Ag_row = sh.max_row + 1
                sh.cell(Ag_row, idx_Desc).value = "AGV go"
                sh.cell(Ag_row, idx_Start).value = nowtime
                sh.cell(Ag_row, idx_WPNo).value = rfid_agv_go[0].wpno
                sh.cell(Ag_row, idx_oNo).value = rfid_agv_go[0].ono
                sh.cell(Ag_row, idx_OpNo).value = rfid_agv_go[0].opno
                sh.cell(Ag_row, idx_oPos).value = rfid_agv_go[0].opos
                del rfid_agv_go[0]
                print("AGV Waiting (go)", rfid_agv_go)
                print("AGV go")
            if M_bg41.get_value() and isAgvGo:
                isAgvGo = False
                sh.cell(Ag_row, idx_end).value = nowtime
                print("OK")
            #回程AGV
            if M_bg21.get_value() and not isRfidReadytoNote_Back:
                isRfidReadytoNote_Back = True
            elif not M_bg21.get_value() and isRfidReadytoNote_Back and M_OpNo_2.get_value() not in [115, 120, 121, 122, 123, 112]:
                isRfidReadytoNote_Back = False
                rfid_agv_back.append(rfid_data(M_WPNo_2.get_value(),M_ONo_2.get_value(),M_oPos_2.get_value(),M_OpNo_2.get_value()))
                print("AGV Waiting (back)", rfid_agv_back)
            if M_bg42.get_value() and not isAgvBackWait:
                isAgvBackWait = True
            elif not M_bg42.get_value() and isAgvBackWait:
                isAgvBackWait = False
                isAgvBack = True
                Ab_row = sh.max_row + 1
                sh.cell(Ab_row, idx_Desc).value = "AGV back"
                sh.cell(Ab_row, idx_Start).value = nowtime
                sh.cell(Ab_row, idx_WPNo).value = rfid_agv_back[0].wpno
                sh.cell(Ab_row, idx_oNo).value = rfid_agv_back[0].ono
                sh.cell(Ab_row, idx_OpNo).value = rfid_agv_back[0].opno
                sh.cell(Ab_row, idx_oPos).value = rfid_agv_back[0].opos
                del rfid_agv_back[0]
                print("AGV Waiting (back)", rfid_agv_back)
                print("AGV back")
            if V_bg41.get_value() and isAgvBack:
                isAgvBack = False
                sh.cell(Ab_row, idx_end).value = nowtime
                print("OK")

            #量測
            if M_mes_busy.get_value() and M_bg51_1.get_value() and not isMeasure:
                isMeasure = True
                M_row = sh.max_row + 1
                sh.cell(M_row, idx_Desc).value = "measure a part (analog)"
                sh.cell(M_row, idx_Start).value = nowtime
                sh.cell(M_row, idx_WPNo).value = M_WPNo_1.get_value()
                sh.cell(M_row, idx_oNo).value = M_ONo_1.get_value()
                sh.cell(M_row, idx_OpNo).value = M_OpNo_1.get_value()
                sh.cell(M_row, idx_oPos).value = M_oPos_1.get_value()
                print("measure a part (analog)")
            elif not M_mes_busy.get_value() and isMeasure:
                isMeasure = False
                sh.cell(M_row, idx_end).value = nowtime
            #鑽孔
            if D_mes_busy.get_value() and D_bg21.get_value() and not isDrill:
                isDrill = True
                D_row = sh.max_row + 1
                sh.cell(D_row, idx_Desc).value = "drilling both"
                sh.cell(D_row, idx_Start).value = nowtime
                sh.cell(D_row, idx_WPNo).value = D_WPNo.get_value()
                sh.cell(D_row, idx_oNo).value = D_ONo.get_value()
                sh.cell(D_row, idx_OpNo).value = D_OpNo.get_value()
                sh.cell(D_row, idx_oPos).value = D_oPos.get_value()
                print("drilling both")
            elif not D_mes_busy.get_value() and isDrill:
                isDrill = False
                sh.cell(D_row, idx_end).value = nowtime
            #PCB放置
            if pcb_mes_busy.get_value() and pcb_bg31.get_value() and not isPcb:
                isPcb = True
                pcb_row = sh.max_row + 1
                sh.cell(pcb_row, idx_Desc).value = "assemble a PCB with both fuses"
                sh.cell(pcb_row, idx_Start).value = nowtime
                sh.cell(pcb_row, idx_WPNo).value = pcb_WPNo.get_value()
                sh.cell(pcb_row, idx_oNo).value = pcb_ONo.get_value()
                sh.cell(pcb_row, idx_OpNo).value = pcb_OpNo.get_value()
                sh.cell(pcb_row, idx_oPos).value = pcb_oPos.get_value()
                print("assemble a PCB with both fuses")
            elif not pcb_mes_busy.get_value() and isPcb:
                isPcb = False
                sh.cell(pcb_row, idx_end).value = nowtime
            #視覺辨識
            if V_mes_busy.get_value() and V_bg51_1.get_value() and not isVaual:
                isVaual = True
                V_row = sh.max_row + 1
                sh.cell(V_row, idx_Desc).value = "check part with camera"
                sh.cell(V_row, idx_Start).value = nowtime
                sh.cell(V_row, idx_WPNo).value = V_WPNo_1.get_value()
                sh.cell(V_row, idx_oNo).value = V_ONo_1.get_value()
                sh.cell(V_row, idx_OpNo).value = V_OpNo_1.get_value()
                sh.cell(V_row, idx_oPos).value = V_oPos_1.get_value()
                print("check part with camera")
            elif not V_mes_busy.get_value() and isVaual:
                isVaual = False
                sh.cell(V_row, idx_end).value = nowtime
            #加蓋
            if C_mes_busy.get_value() and C_bg21.get_value() and not isCover:
                isCover = True
                C_row = sh.max_row + 1
                sh.cell(C_row, idx_Desc).value = "feed back cover from magazine"
                sh.cell(C_row, idx_Start).value = nowtime
                sh.cell(C_row, idx_WPNo).value = C_WPNo.get_value()
                sh.cell(C_row, idx_oNo).value = C_ONo.get_value()
                sh.cell(C_row, idx_OpNo).value = C_OpNo.get_value()
                sh.cell(C_row, idx_oPos).value = C_oPos.get_value()
                print("feed back cover from magazine")
            elif not C_mes_busy.get_value() and isCover:
                isCover = False
                sh.cell(C_row, idx_end).value = nowtime
            #加壓
            if P_mes_busy.get_value() and P_bg21.get_value() and not isPress:
                isPress = True
                P_row = sh.max_row + 1
                sh.cell(P_row, idx_Desc).value = "pressing with force regulation"
                sh.cell(P_row, idx_Start).value = nowtime
                sh.cell(P_row, idx_WPNo).value = P_WPNo.get_value()
                sh.cell(P_row, idx_oNo).value = P_ONo.get_value()
                sh.cell(P_row, idx_OpNo).value = P_OpNo.get_value()
                sh.cell(P_row, idx_oPos).value = P_oPos.get_value()
                print("pressing with force regulation")
            elif not P_mes_busy.get_value() and isPress:
                isPress = False
                sh.cell(P_row, idx_end).value = nowtime
            #加熱
            if H_mes_busy.get_value() and H_bg21.get_value() and not isHeating:
                isHeating = True
                H_row = sh.max_row + 1
                sh.cell(H_row, idx_Desc).value = "heating Part"
                sh.cell(H_row, idx_Start).value = nowtime
                sh.cell(H_row, idx_WPNo).value = H_WPNo.get_value()
                sh.cell(H_row, idx_oNo).value = H_ONo.get_value()
                sh.cell(H_row, idx_OpNo).value = H_OpNo.get_value()
                sh.cell(H_row, idx_oPos).value = H_oPos.get_value()
                sh.cell(H_row, idx_iniTemp).value = H_actual_temp.get_value()
                print("heating Part")
            elif not H_mes_busy.get_value() and isHeating:
                isHeating = False
                sh.cell(H_row, idx_finTemp).value = H_actual_temp.get_value()
                sh.cell(H_row, idx_end).value = nowtime
            #紀錄本次工單oPos最大值
            if C_oPos.get_value() > max_order_number:
                max_order_number = C_oPos.get_value()
    
    except socket.error:
        print_style(style.RED,"連線錯誤，請確認是否已連上Taiwan-CP-Factory後再啟動程式")
    finally:
        excel_save(wb, file_name)
        wb.close()
        H_automatic.set_attribute(ua.AttributeIds.Value, ua.DataValue(True))
        Storage.disconnect()
        Measure.disconnect()
        Drill.disconnect()
        Pcb.disconnect()
        Visual.disconnect()
        Cover.disconnect()
        Press.disconnect()
        Heating.disconnect()

cmd_terminal = threading.Thread(target = terminal)
cmd_terminal.setDaemon(True)
cmd_terminal.start()
main()
