from numpy import index_exp
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import collections
from datetime import datetime

class style():
    """Style Data"""    
    BLACK = '\033[30m'
    RED = '\033[31m'
    GREEN = '\033[32m'
    YELLOW = '\033[33m'
    BLUE = '\033[34m'
    MAGENTA = '\033[35m'
    CYAN = '\033[36m'
    WHITE = '\033[37m'
    BLACK_BG  = '\33[40m'
    RED_BG    = '\33[41m'
    GREEN_BG  = '\33[42m'
    YELLOW_BG = '\33[43m'
    BLUE_BG   = '\33[44m'
    MAGENTA_BG = '\33[45m'
    CYAN_BG  = '\33[46m'
    WHITE_BG  = '\33[47m'
    UNDERLINE = '\033[4m'
    RESET = '\033[0m'

def print_style(Style_in_ClassStyle, str):
    """Print string with different style"""
    print(Style_in_ClassStyle + str + style.RESET)

def SecondCost(start, end):
    """Calculate how many seconds between two datetime"""
    start_detail = [int(datetime.strftime(start,'%d')), # 日
                    int(datetime.strftime(start,'%H')), # 時
                    int(datetime.strftime(start,'%M')), # 分
                    int(datetime.strftime(start,'%S'))] # 秒
    if int(datetime.strftime(start,'%f')) > 500000 : start_detail[3] += 1 #毫秒四捨五入
    end_detail = [int(datetime.strftime(end,'%d')), # 日
                  int(datetime.strftime(end,'%H')), # 時
                  int(datetime.strftime(end,'%M')), # 分
                  int(datetime.strftime(end,'%S'))] # 秒
    if int(datetime.strftime(end,'%f')) > 500000 : end_detail[3] += 1 #毫秒四捨五入
    dt = [end_detail[0]-start_detail[0], 
          end_detail[1]-start_detail[1], 
          end_detail[2]-start_detail[2], 
          end_detail[3]-start_detail[3]]
    duration = ((dt[0]*24 + dt[1])*60 + dt[2])*60 + dt[3]
    return duration

def MachineCheck(desc):
    """Convert processing description to machine number 
    (0:倉儲 1:無人搬運車 2:影像檢測 3:鑽孔 4:PCB放置 5:視覺檢測 6:上蓋放置 7:加壓 8:加熱)"""
    if 'release' in desc: machine = 0
    elif 'store' in desc: machine = 0
    elif 'AGV' in desc: machine = 1
    elif 'measure' in desc: machine = 2
    elif 'drilling' in desc: machine = 3
    elif 'assemble' in desc: machine = 4
    elif 'camera' in desc: machine = 5
    elif 'cover' in desc: machine = 6
    elif 'pressing' in desc: machine = 7
    elif 'heating' in desc: machine = 8
    return machine

def OrderDataQuery(CMD):
    """Search for Work Order Data, including WPNo, oNo, oPos"""
    final_WPNo = []
    final_oNo = []
    final_oPos = []
    final_Start_Time = []
    final_data = [] # final_data[筆數][欄位]
    title = ["WPNo","oNo","oPos","Start-Time"] # 欄位資訊
    title_l = [] # 欄位資訊小寫
    for item in title:
        title_l.append(item.lower())
    output_show = [False]*len(title) # 是否顯示該欄位 
    output_or_not = True
    for i in range(2,sheet.max_row +1):
        WPNo = int(sheet.cell(row=i, column=1).value)
        oNo = int(sheet.cell(row=i, column=3).value)
        oPos = int(sheet.cell(row=i, column=4).value)
        Start_Time = sheet.cell(row=i, column=13).value
        if (oNo not in final_oNo):
            final_oNo.append(oNo)
            final_WPNo.append(WPNo)
            final_oPos.append([])
            final_Start_Time.append([])
        final_oPos[final_oNo.index(oNo)].append(oPos)
        final_Start_Time[final_oNo.index(oNo)].append(Start_Time)
    #儲存資料至 final_data
    try:
        for i in range(len(final_oNo)):
            final_oPos[i] = max(final_oPos[i])
            final_Start_Time[i] = min(final_Start_Time[i])
            final_data.append([final_WPNo[i], final_oNo[i], final_oPos[i], final_Start_Time[i]])
    except IndexError:
        if Console_Language == 'TW':
            print_style(style.RED, "錯誤：資料庫內容不正確。")
        else:
            print_style(style.RED, "Error: Contents in DB is uncorrect.")
    # final_data  = list(zip(final_WPNo, final_oNo, final_oPos, final_Start_Time))
    CMD_s = CMD.split()
    conti = True
    filter_conti = False
    sort_conti = False
    # 顯示判斷
    try:
        index = 0
        while conti:
            target = CMD_s[index]
            conti = False
            if target in title_l:
                output_show[title_l.index(target)] = True
                if "and" in CMD_s[index:]:
                    if CMD_s[index + 1] == "and":
                        index = index + 2
                        conti = True
            elif target == "*" and CMD.count("*") == 1:
                output_show[:] = [True]*len(title)
            else:
                raise Exception
        if len(CMD_s) > index + 1:
            if CMD_s[index + 1] in ["where", "order"]:
                if CMD_s[index + 1] == "where" and len(CMD_s) > index + 2:
                    filter_conti = True
                    index = CMD_s.index("where")
                elif CMD_s[index + 1] == "order" and len(CMD_s) > index + 2:
                    sort_conti = True
                    index = CMD_s.index("order")
                else:
                    raise Exception
            else:
                raise Exception
    except (IndexError,Exception):
        output_or_not = False
        if Console_Language == 'TW':
            print_style(style.RED, "錯誤：指令不存在。")
        else:
            print_style(style.RED, "Error: Command not found.")
    
    # 過濾判斷
    try:
        if filter_conti:
            filter_CMD = CMD_s[index + 1 : ]
            if "order" in CMD_s[index + 1 : ]:
                filter_CMD = CMD_s[index + 1 : CMD_s.index("order")]
                sort_conti = True
                index = CMD_s.index("order")
            for i in range(len(filter_CMD)):
                if filter_CMD[i] == "=" :
                    filter_CMD[i] = "=="
                if filter_CMD[i] == "<>":
                    filter_CMD[i] = "!="
                if filter_CMD[i] == "start-time":
                    filter_CMD[i] = "start_time"
                if "/" in filter_CMD[i] in filter_CMD[i]:
                    try:
                        if ":" in filter_CMD[i]:
                            dateFormatter = "%Y/%m/%d-%H:%M"
                        else:
                            dateFormatter = "%Y/%m/%d"
                        filter_CMD[i] = str(datetime.timestamp(datetime.strptime(filter_CMD[i], dateFormatter)))
                    except:
                        filter_conti = False
                        sort_conti = False
                        output_or_not = False
                        if Console_Language == 'TW':
                            print_style(style.RED, "錯誤：時間格式不正確, 範例：YYYY/mm/dd 或 YYYY/mm/dd-HH:MM")
                        else:
                            print_style(style.RED, "Error: Time format error, ex: YYYY/mm/dd or YYYY/mm/dd-HH:MM")
            filter_CMD = " ".join(filter_CMD)
            count = 0
            for i in range(len(final_data)):
                wpno = final_data[count][0]
                ono = final_data[count][1]
                opos = final_data[count][2]
                start_time = datetime.timestamp(final_data[count][3])
                if not eval(filter_CMD):
                    del final_data[count]
                else:
                    count += 1
    except:
        sort_conti = False
        output_or_not = False
        if Console_Language == 'TW':
            print_style(style.RED, "錯誤：過濾條件不存在, 範例：query * where WPNo == 1211")
        else:
            print_style(style.RED, "Error: Filter condition not found, ex: query * where WPNo == 1211")

    # 排序判斷
    try:
        if sort_conti:
            if CMD_s[index + 1] == "by":
                index = index + 1
            else:
                raise Exception
        while sort_conti:
            sort_conti = False
            target = CMD_s[index + 1]
            operator = CMD_s[index + 2]
            if target in title_l and operator in ['asc','desc']:
                if operator == 'asc':
                    final_data.sort(key=lambda x:x[title_l.index(target)])
                elif operator == 'desc':
                    final_data.sort(key=lambda x:x[title_l.index(target)],reverse=True)
                
                if "and" in CMD_s[index + 1:]:
                    if CMD_s[index + 3] == "and":
                        index = index + 3
                        sort_conti = True
                else:
                    if index + 3 != len(CMD_s):
                        raise Exception
            else:
                raise Exception
    except (IndexError, Exception):
        output_or_not = False
        if Console_Language == 'TW':
            print_style(style.RED, "錯誤：排序條件不存在, 範例：query * order by WPNo desc")
        else:
            print_style(style.RED, "Error: Order condition do not exist. Example: query * order by WPNo desc")
   
    # 輸出
    output_form = ["{:>8}","{:>8}","{:>8}","{:>20}"]
    if output_or_not:
        for i in range(len(title)):
            if output_show[i]:
                print(output_form[i].format(title[i]), end="")
        print("") # 換行
        for i in range(len(final_data)):
            for j in range(len(title)):
                if output_show[j] and j < 3:
                    print(output_form[j].format(final_data[i][j]), end="")
                elif output_show[j] and j == 3:
                    print(output_form[j].format(datetime.strftime(final_data[i][j], '%Y/%m/%d %H:%M')), end="")
            print("") # 換行

def Gannt(oNo_input, Gannt_Language):
    """Draw gantt base on the data in excel."""
    task_data = collections.namedtuple('task_data', 'machine, oPos, Start_Time, duration')
    final_data = {}  # 儲存排序後結果
    bar_color = ['red', 'blue', 'green', 'gold', 'chocolate', 'gray', 'blueviolet', 'hotpink', 'darkred', 'black'] # 甘特圖工單代表顏色
    machine_name_en = ["Storage", 'AGV', 'Check part', 'Milling', 'PCB', 'Fuse check', 'Cover', 'Pressure', 'Heat']
    machine_name_tw = ['倉儲　　　', '無人搬運車', '影像檢測　', '鑽孔　　　', 'PCB放置 　', '視覺檢測　', '上蓋放置　', '加壓　　　', '加熱　　　']
    final_oNo = [] 
    start_dt = [] # 以Datetime格式儲存起始時間
    oPos_Max = 0
    total_time_dt = datetime(2000, 1, 1, 0, 0, 0)
    for i in range(2,sheet.max_row +1):
        oNo = int(sheet.cell(row=i, column=3).value)
        if (oNo not in final_oNo): final_oNo.append(oNo)
        if oNo == oNo_input:
            oPos = int(sheet.cell(row=i, column=4).value)
            start_dt.append(sheet.cell(row=i, column=13).value)
            if (int(sheet.cell(row=i, column=4).value) > oPos_Max): oPos_Max = int(sheet.cell(row=i, column=4).value)
            if (sheet.cell(row=i, column=14).value > total_time_dt): total_time_dt = sheet.cell(row=i, column=14).value
    if oNo_input in final_oNo:
        time0 = min(start_dt)
        total_time = SecondCost(time0, total_time_dt)
        where_order_at = ['A'] * oPos_Max # 記錄黑色小車在哪一區，用來判斷AGV去或回
        for i in range(2,sheet.max_row +1):
            oNo = int(sheet.cell(row=i, column=3).value)
            oPos = int(sheet.cell(row=i, column=4).value)
            machine = MachineCheck(sheet.cell(row=i, column=5).value)
            Start_Time = SecondCost(time0, sheet.cell(row=i, column=13).value)
            duration = SecondCost(sheet.cell(row=i, column=13).value, sheet.cell(row=i, column=14).value)
            if oNo == oNo_input:
                final_data[i] = task_data(machine, oPos, Start_Time, duration)
                plt.barh(final_data[i].machine, final_data[i].duration, 
                        left = final_data[i].Start_Time,
                        color = bar_color[final_data[i].oPos -1],
                        height = 0.9)
                # ---判斷AGV去或回並顯示在甘特圖---
                # Go: A->B，Back: B->A
                if final_data[i].machine == 1:
                    if where_order_at[oPos -1] == 'A':
                        bar_text = "Go"
                        where_order_at[oPos -1] = 'B'
                    elif where_order_at[oPos -1] == 'B':
                        bar_text = "Back"
                        where_order_at[oPos -1] = 'A'
                    plt.text(final_data[i].Start_Time, 1, bar_text,color="white")
                # --------------------------------        
        # 圖表多語系設定
        if Gannt_Language == 'TW':
            for i in range(oPos_Max):
                plt.barh(0,0,label = '工單 ' + str(i +1),color = bar_color[i])
            plt.xlabel('時間 (分鐘)')
            plt.title('甘特圖 ( oNo: ' + str(oNo_input) + '，總時間: ' + str(total_time) + ' 秒 )')
            plt.yticks(range(len(machine_name_tw)), machine_name_tw)
            plt.rcParams['font.sans-serif'] = ['Microsoft JhengHei'] #甘特圖中文字型設定
            plt.rcParams['axes.unicode_minus'] = False
        else:
            for i in range(oPos_Max):
                plt.barh(0,0,label = 'Order ' + str(i +1),color = bar_color[i])
            plt.xlabel('Time (min.)')
            plt.title('Gannt ( oNo: ' + str(oNo_input) + ', Time Cost: ' + str(total_time) + ' seconds )')
            plt.yticks(range(len(machine_name_en)), machine_name_en)
            
        plt.xticks(range(0, int(total_time) +60, 60), range(int(total_time/60)+2))
        plt.xticks(range(0, int(total_time) +60, 30))
        plt.grid(True, color = "grey", linewidth = "0.5", linestyle='dashed', axis = 'x') # 圖表格線設定
        plt.legend(loc = 'upper left')
        plt.show()    
    else:
        if Console_Language == 'TW':
            print_style(style.RED, "錯誤：工單編號 " + str(oNo_input) + " 不存在於資料庫中。")
        else:
            print_style(style.RED, "Error: Work Order Number " + str(oNo_input) + " does not exist in the database.")

def Console():
    global Console_Language, wb, sheet
    while True:
        command = input(">> ").lower()
        if command == "help" or command == "-h":
            if Console_Language == 'TW':
                print("指令資訊:")
                print(" 輸入 'show'+'工單編號(oNo)' 顯示該工單的甘特圖, 例: 'show 1816'。")
                print("  若甘特圖無法顯示中文，可在指令最後加上 'en' 來顯示英文甘特圖，例: 'show 1816 en'。")
                print(" 輸入 'query *' 查詢資料庫內所有工單資訊。")
                print(" 輸入 'refresh' 重新讀取資料庫檔案。")
                print(" 輸入 'tw' 或 'en' 切換終端機顯示語言。")
                print(" 輸入 'exit' 結束程式。")
            else:
                print("Command Information:")
                print(" Enter 'show'+'oNo' to show gannt of the work order number, e.g. 'show 1816'.")
                print(" Enter 'query *' to query about all the orders information in the database.")
                print(" Enter 'refresh' to reload the database file.")
                print(" Enter 'tw' or 'en' to switch the language used in cosole.")
                print(" Enter 'exit' to leave the program.")        
        elif command.startswith('show') and len(command.split()) <= 3:
            show = command.split()
            if show[0] == 'show' and len(show) == 1:
                if Console_Language == 'TW':
                    print_style(style.RED, "錯誤：請輸入 '單號(oNo)' 在 'show' 後方")
                else:
                    print_style(style.RED, "Error: Please input 'oNo' after 'show'.")
            elif show[0] == 'show' and len(show) == 2:
                try:
                    Gannt(int(show[1]),Console_Language)
                except ValueError:
                    if Console_Language == 'TW':
                        print_style(style.RED, "錯誤：'單號(oNo)' 必須為整數")
                    else:
                        print_style(style.RED, "Error: 'oNo' should be a integer.")
            elif show[0] == 'show' and len(show) == 3 and show[2] == 'en':
                try:
                    Gannt(int(show[1]), show[2].upper)
                except ValueError:
                    if Console_Language == 'TW':
                        print_style(style.RED, "錯誤：'單號(oNo)' 必須為整數")
                    else:
                        print_style(style.RED, "Error: 'oNo' should be a integer.")
            else:
                if Console_Language == 'TW':
                    print_style(style.RED, "錯誤：指令不存在。")
                else:
                    print_style(style.RED, "Error: Command not found.")
        elif command.startswith('query'):
            if command[5:] == " help" or command[5:] == " -h":
                if Console_Language == 'TW':
                    print("搜尋詳細指令資訊:")
                    print(" 架構：query (1) where (2) order by (3)")
                    print(" (1) 選擇顯示項目，例: query WPNo and oNo")
                    print(" (2) 欄位條件，例: query * where WPNo = 1211")
                    print("     若輸入時間，格式須為: YYYY/mm/dd 或 YYYY/mm/dd-HH:MM")
                    print(" (3) 排序條件，例: query * order by WPNo asc")
                else:
                    print("Details of query commands:")
                    print(" Structure: query (1) where (2) order by (3)")
                    print(" (1) Rows to show, ex: query WPNo and oNo")
                    print(" (2) Filter for columns, ex: query * where WPNo = 1211")
                    print("     For time format, it must be YYYY/mm/dd or YYYY/mm/dd-HH:MM")
                    print(" (3) Sort condition, ex: query * order by WPNo asc")
            else:
                OrderDataQuery(command[5:])
        elif command == "tw" or command == "en":
            Console_Language = command.upper()
            if Console_Language == 'TW':
                print_style(style.CYAN, "歡迎使用甘特圖繪製，輸入指令以繼續。或輸入'help'取得指令說明")
            else:
                print_style(style.CYAN, "Welcome to gannt drawing, enter command to continue. Or enter 'help' for more infomation")
        elif command == "refresh":
            try:
                wb = load_workbook('tblFinStep.xlsx')
                sheet = wb['tblFinStep']
                if Console_Language == 'TW':
                    print("資料庫更新成功。")
                else:
                    print("Refresh Succeeded.")
            except:
                if Console_Language == 'TW':
                    print_style(style.RED, "錯誤：更新失敗。")
                else:
                    print_style(style.RED, "Error: Refresh failed.")
        elif command == "exit":
            if Console_Language == 'TW':
                exit("感謝使用！")
            else:
                exit("Thanks for using.")
        else:
            if Console_Language == 'TW':
                print_style(style.RED, "錯誤：指令不存在。")
            else:
                print_style(style.RED, "Error: Command not found.")

Console_Language = 'TW' # 語言設定 'TW' for Tranditional Chinese, 'EN' for English
try:
    wb = load_workbook('tblFinStep.xlsx')
    sheet = wb['tblFinStep']
    if Console_Language == 'TW':
        print_style(style.CYAN, "歡迎使用甘特圖繪製，輸入指令以繼續。或輸入'help'取得指令說明")
        print("資料庫讀取成功。")
    else:
        print_style(style.CYAN, "Welcome to gannt drawing, enter command to continue. Or enter 'help' for more infomation")
        print("Database load succeeded.")
except:
    if Console_Language == 'TW':
        print_style(style.RED, "錯誤：資料庫讀取失敗。")
    else:
        print_style(style.RED, "Error: Database load failed.")
    exit()
Console()